import uuid

from django.conf import settings

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from .models import Device


def autofit_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:  # noqa
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width


def create_excel(ids):
    """
        Export data to Excel file

    Args:
        ids (list): list of Device id

    Returns:
        url: str excel file url
    """
    wb = Workbook()
    ws = wb.active
    # styles
    bold = Font(bold=True)
    thin = Side(border_style="thin", color="000000")
    fill = PatternFill("solid", fgColor="00CCFFCC")
    # table head
    ws['A1'] = '№'
    ws['B1'] = 'IP адрес'
    ws['C1'] = 'Тип устройства'
    ws['D1'] = 'Адрес'
    ws['E1'] = 'Кабинет'
    ws['F1'] = 'Служба'
    ws['G1'] = 'Подразделение'
    ws['H1'] = 'Описание'
    ws['I1'] = 'Сотрудник'
    # font, color and color
    for r in ws["A1:I1"]:
        for cell in r:
            cell.font = bold
            cell.fill = fill
            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )
    # table body
    row = 2
    number = 1
    for id in ids:
        try:
            device = Device.objects.get(id=id)
            ws[f'A{row}'] = number
            ws[f'B{row}'] = device.ip
            ws[f'C{row}'] = device.type_of_device.title \
                if device.type_of_device else None
            ws[f'D{row}'] = device.location.title \
                if device.location else None
            ws[f'E{row}'] = device.cabinet
            ws[f'F{row}'] = device.department.title \
                if device.department else None
            ws[f'G{row}'] = device.management.title \
                if device.management else None
            ws[f'H{row}'] = device.description
            ws[f'I{row}'] = device.employee
            # border
            for r in ws[f'A{row}:I{row}']:
                for cell in r:
                    cell.border = Border(left=thin, right=thin,
                                         top=thin, bottom=thin)
            row = row + 1
            number = number + 1
        except Exception:
            ws[f'A{row}'] = number

    # autofit columns
    autofit_columns(ws)
    # saving to excel file
    filename = f'{uuid.uuid4()}.xlsx'
    wb.save(filename=f"{settings.MEDIA_ROOT}/{filename}")

    return f"{settings.MEDIA_URL}{filename}"
